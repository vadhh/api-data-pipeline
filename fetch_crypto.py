# fetch_crypto.py

import time
import sys
import requests
import os
import math
import argparse
import logging
from datetime import date
from openpyxl import Workbook, load_workbook

# Configure logging to write to both stdout and pipeline.log
logger = logging.getLogger("crypto_pipeline")
logger.setLevel(logging.INFO)

if not logger.handlers:
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    
    # File handler
    fh = logging.FileHandler("pipeline.log", encoding="utf-8")
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    
    # Stream handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(formatter)
    logger.addHandler(ch)


HEADERS = [
    "Rank",
    "Name",
    "Symbol",
    "Current Price (USD)",
    "24h Change %",
    "7d Change %",
    "Market Cap",
    "24h Volume",
    "Circulating Supply",
    "Volatility",
]


def derive_volatility(change_24h):
    if change_24h is None:
        return "NORMAL"
    return "HIGH" if abs(change_24h) > 10 else "NORMAL"


def transform_coin(raw):
    change_24h = raw.get("price_change_percentage_24h")
    return {
        "Rank": raw.get("market_cap_rank"),
        "Name": raw.get("name"),
        "Symbol": raw.get("symbol", "").upper() if raw.get("symbol") is not None else "",
        "Current Price (USD)": raw.get("current_price"),
        "24h Change %": change_24h,
        "7d Change %": raw.get("price_change_percentage_7d_in_currency"),
        "Market Cap": raw.get("market_cap"),
        "24h Volume": raw.get("total_volume"),
        "Circulating Supply": raw.get("circulating_supply"),
        "Volatility": derive_volatility(change_24h),
    }


COINGECKO_URL = "https://api.coingecko.com/api/v3/coins/markets"

PARAMS_BASE = {
    "vs_currency": "usd",
    "order": "market_cap_desc",
    "per_page": 100,
    "sparkline": "false",
    "price_change_percentage": "7d",
}


def fetch_page(page, max_retries=5):
    params = {**PARAMS_BASE, "page": page}
    for attempt in range(max_retries):
        try:
            resp = requests.get(COINGECKO_URL, params=params, timeout=30)
            if resp.status_code == 429:
                wait = 15 * (attempt + 1)
                logger.warning(f"  Rate limit (429) hit. Retry {attempt + 1}/{max_retries} for page {page} "
                               f"(waiting {wait}s)...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            wait = 5 * (2 ** attempt)
            logger.warning(f"  Retry {attempt + 1}/{max_retries} for page {page} "
                           f"(waiting {wait}s): {e}")
            time.sleep(wait)
    return None


def fetch_all_coins(top=250):
    all_coins = []
    per_page = PARAMS_BASE.get("per_page", 100)
    pages_needed = math.ceil(top / per_page)
    for page in range(1, pages_needed + 1):
        logger.info(f"Fetching page {page}/{pages_needed}...")
        data = fetch_page(page)
        if data is None:
            logger.error(f"ERROR: Failed to fetch page {page} after retries. Aborting.")
            sys.exit(1)
        all_coins.extend(data)
        if page < pages_needed:
            time.sleep(2)
    all_coins = all_coins[:top]
    logger.info(f"Fetched {len(all_coins)} coins.")
    return [transform_coin(c) for c in all_coins]


def get_sheet_name(wb, base_name):
    if base_name not in wb.sheetnames:
        return base_name
    counter = 2
    while f"{base_name}_{counter}" in wb.sheetnames:
        counter += 1
    return f"{base_name}_{counter}"


def write_to_excel(rows, filepath, sheet_date=None):
    if sheet_date is None:
        sheet_date = date.today().isoformat()

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = get_sheet_name(wb, sheet_date)
    ws = wb.create_sheet(title=sheet_name)

    bold = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, coin in enumerate(rows, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=coin.get(header))

    col_widths = {
        "Rank": 6,
        "Name": 22,
        "Symbol": 8,
        "Current Price (USD)": 20,
        "24h Change %": 14,
        "7d Change %": 14,
        "Market Cap": 18,
        "24h Volume": 16,
        "Circulating Supply": 20,
        "Volatility": 12,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(header, 14)

    ws.freeze_panes = "A2"

    wb.save(filepath)
    wb.close()
    logger.info(f"Saved sheet '{sheet_name}' to {filepath}")


def main():
    parser = argparse.ArgumentParser(description="Crypto Data Pipeline")
    parser.add_argument("--top", type=int, default=250, help="Number of top coins to fetch")
    parser.add_argument("--output", type=str, default="crypto_data.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    logger.info("=" * 50)
    logger.info("Crypto Data Pipeline")
    logger.info(f"Date: {date.today().isoformat()}")
    logger.info(f"Target: Top {args.top} coins to {args.output}")
    logger.info("=" * 50)

    coins = fetch_all_coins(top=args.top)
    write_to_excel(coins, args.output)

    logger.info("=" * 50)
    logger.info("Done!")


if __name__ == "__main__":
    main()
