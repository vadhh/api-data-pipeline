# tests/test_excel.py

import os
from openpyxl import load_workbook
from fetch_crypto import write_to_excel, HEADERS


SAMPLE_ROWS = [
    {
        "Rank": 1,
        "Name": "Bitcoin",
        "Symbol": "BTC",
        "Current Price (USD)": 68432.12,
        "24h Change %": 2.5,
        "7d Change %": -1.3,
        "Market Cap": 1340000000000,
        "24h Volume": 28000000000,
        "Circulating Supply": 19700000.0,
        "Volatility": "NORMAL",
    },
    {
        "Rank": 2,
        "Name": "Ethereum",
        "Symbol": "ETH",
        "Current Price (USD)": 3520.45,
        "24h Change %": -11.2,
        "7d Change %": 4.8,
        "Market Cap": 423000000000,
        "24h Volume": 15000000000,
        "Circulating Supply": 120000000.0,
        "Volatility": "HIGH",
    },
]


def test_creates_new_workbook(tmp_path):
    path = str(tmp_path / "test.xlsx")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-27")

    wb = load_workbook(path)
    assert "2026-05-27" in wb.sheetnames
    ws = wb["2026-05-27"]
    headers = [cell.value for cell in ws[1]]
    assert headers == HEADERS
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Bitcoin"
    assert ws.cell(row=3, column=10).value == "HIGH"
    wb.close()


def test_appends_sheet_to_existing(tmp_path):
    path = str(tmp_path / "test.xlsx")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-27")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-28")

    wb = load_workbook(path)
    assert "2026-05-27" in wb.sheetnames
    assert "2026-05-28" in wb.sheetnames
    wb.close()


def test_duplicate_date_gets_suffix(tmp_path):
    path = str(tmp_path / "test.xlsx")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-27")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-27")

    wb = load_workbook(path)
    assert "2026-05-27" in wb.sheetnames
    assert "2026-05-27_2" in wb.sheetnames
    wb.close()


def test_header_row_is_bold(tmp_path):
    path = str(tmp_path / "test.xlsx")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-27")

    wb = load_workbook(path)
    ws = wb["2026-05-27"]
    for cell in ws[1]:
        assert cell.font.bold is True
    wb.close()


def test_freeze_panes(tmp_path):
    path = str(tmp_path / "test.xlsx")
    write_to_excel(SAMPLE_ROWS, path, "2026-05-27")

    wb = load_workbook(path)
    ws = wb["2026-05-27"]
    assert ws.freeze_panes == "A2"
    wb.close()
